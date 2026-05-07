from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

wb = Workbook()

# 配色方案
COLOR = {
    'primary': '065A82',      # 深蓝
    'secondary': '1C7293',    # 青蓝
    'accent': '21295C',       # 深夜蓝
    'light': 'CADCFC',        # 冰蓝
    'success': '10B981',      # 绿色
    'warning': 'F59E0B',      # 橙色
    'error': 'EF4444',        # 红色
    'bg': 'F8FAFC'           # 浅灰背景
}

# 定义边框样式
def get_border(style='thin'):
    return Border(
        left=Side(style=style),
        right=Side(style=style),
        top=Side(style=style),
        bottom=Side(style=style)
    )

# ==================== 工作表1：流程图 ====================
ws1 = wb.active
ws1.title = "物料流转流程图"

# 设置列宽
for col in range(1, 15):
    ws1.column_dimensions[get_column_letter(col)].width = 15

# 标题
ws1['A1'] = '物料流转流程图'
ws1['A1'].font = Font(name='Microsoft YaHei', size=16, bold=True, color='FFFFFF')
ws1['A1'].fill = PatternFill(start_color=COLOR['primary'], end_color=COLOR['primary'], fill_type='solid')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.merge_cells('A1:N1')
ws1.row_dimensions[1].height = 35

# 流程步骤数据
flow_steps = [
    {'step': 1, 'name': '仓库原料收料', 'role': '仓库', 'desc': '接收供应商来料'},
    {'step': 2, 'name': 'IQC检验', 'role': 'IQC', 'desc': '进料质量检验'},
    {'step': 3, 'name': '仓库收料并\n对生产发料', 'role': '仓库', 'desc': '入库存储并配送至产线'},
    {'step': 4, 'name': '生产制造', 'role': '生产', 'desc': '按工艺要求生产'},
    {'step': 5, 'name': '过程半成品检验', 'role': '生产全检', 'desc': '制程中质量检查'},
    {'step': 6, 'name': '半成品入库', 'role': '仓库', 'desc': '合格半成品入库'},
    {'step': 7, 'name': '半成品外发加工', 'role': '生管', 'desc': '安排委外加工'},
    {'step': 8, 'name': '厂商加工', 'role': '生管', 'desc': '委外厂商加工'},
    {'step': 9, 'name': '委外成品回厂', 'role': '成品委外仓库', 'desc': '加工完成回厂接收'},
    {'step': 10, 'name': '成品回厂检验', 'role': 'IQC', 'desc': '委外成品入厂检验'},
    {'step': 11, 'name': '成品入库', 'role': '仓库', 'desc': '合格成品入库存储'},
    {'step': 12, 'name': '成品出库全检', 'role': '仓库/生产全检', 'desc': '出货前全检'},
    {'step': 13, 'name': '成品全检入库', 'role': '生产全检', 'desc': '全检后重新入库'},
    {'step': 14, 'name': '成品检验', 'role': 'IQC', 'desc': '最终出货检验'},
    {'step': 15, 'name': '成品出货', 'role': '仓库/生管', 'desc': '安排出货给客户'}
]

# 创建流程图（使用单元格模拟流程图）
row = 3
for i, step in enumerate(flow_steps):
    # 步骤编号
    ws1[f'A{row}'] = step['step']
    ws1[f'A{row}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
    ws1[f'A{row}'].fill = PatternFill(start_color=COLOR['primary'], end_color=COLOR['primary'], fill_type='solid')
    ws1[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws1[f'A{row}'].border = get_border()
    
    # 步骤名称
    ws1[f'B{row}'] = step['name']
    ws1[f'B{row}'].font = Font(name='Microsoft YaHei', size=11, bold=True)
    ws1[f'B{row}'].fill = PatternFill(start_color=COLOR['light'], end_color=COLOR['light'], fill_type='solid')
    ws1[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1[f'B{row}'].border = get_border()
    ws1.merge_cells(f'B{row}:D{row}')
    
    # 负责岗位
    ws1[f'E{row}'] = step['role']
    ws1[f'E{row}'].font = Font(name='Microsoft YaHei', size=11, bold=True, color=COLOR['secondary'])
    ws1[f'E{row}'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    ws1[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws1[f'E{row}'].border = get_border()
    ws1.merge_cells(f'E{row}:G{row}')
    
    # 说明
    ws1[f'H{row}'] = step['desc']
    ws1[f'H{row}'].font = Font(name='Microsoft YaHei', size=10, color='595959')
    ws1[f'H{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws1[f'H{row}'].border = get_border()
    ws1.merge_cells(f'H{row}:N{row}')
    
    # 箭头（除了最后一步）
    if i < len(flow_steps) - 1:
        ws1[f'A{row+1}'] = '↓'
        ws1[f'A{row+1}'].font = Font(name='Microsoft YaHei', size=16, color=COLOR['secondary'], bold=True)
        ws1[f'A{row+1}'].alignment = Alignment(horizontal='center', vertical='center')
        ws1.merge_cells(f'A{row+1}:N{row+1}')
        row += 1
    
    row += 1
    ws1.row_dimensions[row-1].height = 35

# ==================== 工作表2：职责矩阵 ====================
ws2 = wb.create_sheet("岗位职责矩阵")

# 设置列宽
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 12

# 标题
ws2['A1'] = '岗位职责矩阵'
ws2['A1'].font = Font(name='Microsoft YaHei', size=16, bold=True, color='FFFFFF')
ws2['A1'].fill = PatternFill(start_color=COLOR['primary'], end_color=COLOR['primary'], fill_type='solid')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.merge_cells('A1:H1')
ws2.row_dimensions[1].height = 35

# 表头
headers = ['序号', '流程步骤', '仓库', 'IQC', '生产', '生管', '成品委外仓库', '生产全检']
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=2, column=col, value=header)
    cell.font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR['secondary'], end_color=COLOR['secondary'], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = get_border()

ws2.row_dimensions[2].height = 40

# 流程数据
process_data = [
    (1, '仓库原料收料', '●', '', '', '', '', ''),
    (2, 'IQC检验', '', '●', '', '', '', ''),
    (3, '仓库收料并对生产发料', '●', '', '', '', '', ''),
    (4, '生产制造', '', '', '●', '', '', ''),
    (5, '过程半成品检验', '', '', '■', '', '', ''),  # ■表示生产全检参与
    (6, '半成品入库', '●', '', '', '', '', ''),
    (7, '半成品外发加工', '', '', '', '●', '', ''),
    (8, '厂商加工', '', '', '', '●', '', ''),
    (9, '委外成品回厂', '', '', '', '', '●', ''),
    (10, '成品回厂检验', '', '■', '', '', '●', ''),  # ■表示IQC参与
    (11, '成品入库', '●', '', '', '', '', ''),
    (12, '成品出库全检', '●', '', '', '', '', '●'),
    (13, '成品全检入库', '●', '', '', '', '', '●'),
    (14, '成品检验', '', '■', '', '', '', '●'),  # ■表示IQC参与
    (15, '成品出货', '●', '', '', '●', '', '')  # 生管也参与
]

# 填充数据
for idx, (seq, step, wh, iqc, prod, pc, out_wh, prod_qc) in enumerate(process_data, 3):
    # 序号
    cell = ws2.cell(row=idx, column=1, value=seq)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = get_border()
    cell.font = Font(name='Microsoft YaHei', size=10, bold=True)
    
    # 流程步骤
    cell = ws2.cell(row=idx, column=2, value=step)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = get_border()
    cell.font = Font(name='Microsoft YaHei', size=10)
    
    # 各岗位职责
    roles = [wh, iqc, prod, pc, out_wh, prod_qc]
    for col_offset, role in enumerate(roles, 3):
        cell = ws2.cell(row=idx, column=col_offset, value=role if role else '')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = get_border()
        cell.font = Font(name='Microsoft YaHei', size=14)
        
        # 设置颜色
        if role == '●':
            cell.font = Font(name='Microsoft YaHei', size=14, color=COLOR['primary'], bold=True)
            cell.fill = PatternFill(start_color=COLOR['light'], end_color=COLOR['light'], fill_type='solid')
        elif role == '■':
            cell.font = Font(name='Microsoft YaHei', size=14, color=COLOR['warning'], bold=True)
            cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    
    ws2.row_dimensions[idx].height = 25

# 添加图例说明
legend_row = len(process_data) + 4
ws2[f'A{legend_row}'] = '图例说明：'
ws2[f'A{legend_row}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color=COLOR['primary'])

ws2[f'A{legend_row+1}'] = '●'
ws2[f'A{legend_row+1}'].font = Font(name='Microsoft YaHei', size=16, color=COLOR['primary'], bold=True)
ws2[f'B{legend_row+1}'] = '主要责任'
ws2[f'B{legend_row+1}'].font = Font(name='Microsoft YaHei', size=11)

ws2[f'A{legend_row+2}'] = '■'
ws2[f'A{legend_row+2}'].font = Font(name='Microsoft YaHei', size=16, color=COLOR['warning'], bold=True)
ws2[f'B{legend_row+2}'] = '协作/参与'
ws2[f'B{legend_row+2}'].font = Font(name='Microsoft YaHei', size=11)

# 添加流程说明
desc_row = legend_row + 4
ws2[f'A{desc_row}'] = '岗位职责说明：'
ws2[f'A{desc_row}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color=COLOR['primary'])

descriptions = [
    '仓库：负责原料接收、存储、发放和成品出入库管理',
    'IQC：负责进料检验、委外成品回厂检验和成品出货检验',
    '生产：负责按工艺要求进行生产制造',
    '生产全检：负责过程半成品检验、成品出库全检和入库全检',
    '生管：负责生产计划安排、委外加工协调和出货安排',
    '成品委外仓库：负责委外加工物料的接收和管理'
]

for i, desc in enumerate(descriptions, desc_row + 1):
    ws2[f'A{i}'] = desc
    ws2[f'A{i}'].font = Font(name='Microsoft YaHei', size=10)
    ws2.merge_cells(f'A{i}:H{i}')

# ==================== 保存文件 ====================
output_file = '物料流转流程与岗位职责_v2.xlsx'
wb.save(output_file)
print(f"✅ Excel文件生成成功：{output_file}")
print(f"📊 包含2个工作表：")
print(f"   1. 物料流转流程图 - 可视化流程展示")
print(f"   2. 岗位职责矩阵 - 详细职责标注")
print(f"📋 共{len(flow_steps)}个流程步骤")
print(f"👥 共6个岗位职责标注")
