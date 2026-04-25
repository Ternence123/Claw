import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

wb = Workbook()

# ── 色板 ──────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT_RED = "C00000"
ACCENT_ORG = "ED7D31"
ACCENT_GRN = "70AD47"
GOLD       = "FFD700"
WHITE      = "FFFFFF"
GRAY_HDR   = "D9D9D9"
LIGHT_GRAY = "F5F5F5"
MED_GRAY   = "BFBFBF"

def fill(hex_):  return PatternFill("solid", fgColor=hex_)
def font(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_border():
    t = Side(style="medium")
    th = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)

def hdr(ws, row, col, text, bg=MID_BLUE, fg=WHITE, bold=True, size=10,
        h="center", wrap=True, colspan=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font  = font(bold=bold, color=fg, size=size)
    c.fill  = fill(bg)
    c.alignment = align(h=h, v="center", wrap=wrap)
    c.border = border()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)
    return c

def cell(ws, row, col, val, bold=False, bg=None, fg="000000",
         h="left", wrap=True, num_fmt=None, border_style="thin"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = font(bold=bold, color=fg, size=10)
    c.alignment = align(h=h, v="center", wrap=wrap)
    c.border    = border(style=border_style)
    if bg:
        c.fill = fill(bg)
    if num_fmt:
        c.number_format = num_fmt
    return c

# ══════════════════════════════════════════════════════
# SHEET 1 — 品牌对比总表
# ══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "品牌对比总表"
ws1.sheet_view.showGridLines = False

# 列宽
col_widths = [6, 16, 10, 15, 8, 8, 8, 8, 8, 8, 8, 8, 10, 10, 10, 10, 10, 13]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.row_dimensions[1].height = 18
ws1.row_dimensions[2].height = 18
ws1.row_dimensions[3].height = 36

# ── 标题行 ───────────────────────────────────────────
ws1.merge_cells("A1:R1")
t = ws1["A1"]
t.value = "MIM行业 ERP系统 品牌对比分析总表"
t.font  = Font(bold=True, color=WHITE, size=14, name="Arial")
t.fill  = fill(DARK_BLUE)
t.alignment = align(h="center", v="center")
t.border = border(style="medium")

ws1.merge_cells("A2:R2")
sub = ws1["A2"]
sub.value = "数据来源：IT之家、腾讯新闻、厂商官网 | 整理时间：2026年4月 | ⚠️ 价格仅供参考，实际以厂商报价为准"
sub.font  = Font(color="595959", size=9, name="Arial", italic=True)
sub.fill  = fill(GRAY_HDR)
sub.alignment = align(h="center", v="center")

# ── 表头 ─────────────────────────────────────────────
hdrs = [
    "序号", "品牌/产品", "厂商来源", "目标客群",
    "模具管理", "工艺路线", "物料追溯", "质量管理",
    "生产排程", "设备集成", "财务模块", "移动端",
    "云/SaaS", "AI功能", "实施周期", "典型价格区间",
    "MIM适配度", "综合评级"
]
for c, h_text in enumerate(hdrs, 1):
    hdr(ws1, 3, c, h_text)

# ── 数据 ─────────────────────────────────────────────
data = [
    # 序号 品牌  来源  目标客群 模具 工艺 物料 质量 排程 设备 财务 移动 云  AI 周期 价格         适配度 综合
    (1,"鼎捷E10","国产","中大型MIM企业\n（年营收1-5亿）",5,5,5,5,5,4,5,4,4,4,"60-90天","40-100万",5,5),
    (2,"鼎捷T100","国产","集团型MIM企业\n（年营收5亿+）",5,5,5,5,5,5,5,5,5,5,"90-180天","100万+",5,5),
    (3,"鼎捷易飞","国产","中小型MIM企业\n（年营收<1亿）",4,4,4,4,4,3,4,4,4,3,"15-30天","15-40万",4,4),
    (4,"用友U9 Cloud","国产","中大型离散制造\n（含精密MIM）",4,5,4,5,5,4,5,5,5,4,"45-60天","25-80万/年",4,4),
    (5,"用友BIP","国产","大型集团企业",4,5,4,5,5,5,5,5,5,5,"60-120天","80万+/年",4,4),
    (6,"金蝶云·星空","国产","中小制造企业\n（MIM精密件）",4,4,4,4,4,3,5,5,4,3,"30-60天","8-25万/年",3,3),
    (7,"金蝶云·星辰","国产","小微制造企业",3,3,3,3,3,2,4,4,3,2,"15-30天","数千元/年",2,2),
    (8,"SAP S/4HANA","外资","跨国大型制造集团",5,5,5,5,5,5,5,5,5,5,"90-180天","80万+/模块\n千万级总投入",5,5),
    (9,"Oracle Fusion","外资","全球化运营企业",5,5,5,5,5,5,5,5,5,5,"90-180天","80万+/模块",5,5),
    (10,"浪潮GS Cloud","国产","大型集团与央国企",4,4,5,4,4,4,5,4,5,4,"60-120天","50-200万",4,4),
    (11,"华为云ERP","国产","有数字化转型需求\n（国产化替代）",4,4,4,4,4,5,5,4,5,5,"45-90天","30-100万/年",4,4),
    (12,"微软Dynamics 365","外资","已使用Microsoft生态\n的企业",4,4,4,4,4,4,4,5,5,5,"60-120天","50-150万/年",3,3),
    (13,"赛意信息SMOM","国产","精密制造/汽车零部件\n（MIM相关）",5,5,5,5,4,5,4,4,4,3,"30-60天","30-80万",4,4),
    (14,"Infor ERP","外资","机械/电子等\n离散制造垂直行业",4,4,4,4,4,4,4,4,4,4,"60-120天","40-100万",4,3),
    (15,"QAD ERP","外资","汽车零部件/航空\n（精益MIM）",4,5,5,5,5,4,4,3,4,3,"60-120天","50-120万",4,3),
]

score_fill = {
    5: ("C6EFCE", "375623"),   # 深绿
    4: ("E2EFDA", "4E7C3B"),   # 浅绿
    3: ("FFEB9C", "7D6608"),   # 黄
    2: ("FFCC99", "BF4B00"),   # 橙
    1: ("FFC7CE", "9C0006"),   # 红
}

def score_color(v):
    try:
        v = int(v)
    except:
        return None, None
    if v >= 5: return "C6EFCE", "375623"
    if v >= 4: return "E2EFDA", "4E7C3B"
    if v >= 3: return "FFEB9C", "7D6608"
    if v >= 2: return "FFCC99", "BF4B00"
    return "FFC7CE", "9C0006"

alt_rows = ["F8FBFF", "FFFFFF"]
for r_idx, row in enumerate(data, 4):
    bg_row = alt_rows[r_idx % 2]
    ws1.row_dimensions[r_idx].height = 40

    for c_idx, val in enumerate(row, 1):
        if c_idx <= 4 or c_idx == 16 or c_idx == 17:
            # 文字列
            bg = bg_row
            c = cell(ws1, r_idx, c_idx, val, bg=bg, h="left", wrap=True)
        elif c_idx == 18:
            # 综合评级：星级文字
            bg, fg_c = score_color(val)
            c = cell(ws1, r_idx, c_idx, val, bold=True, bg=bg, fg=fg_c,
                     h="center")
        elif c_idx in (5,6,7,8,9,10,11,12,13,14,15):
            bg, fg_c = score_color(val)
            c = cell(ws1, r_idx, c_idx, val, bold=True, bg=bg, fg=fg_c,
                     h="center")
        else:
            c = cell(ws1, r_idx, c_idx, val, h="center", bg=bg_row, wrap=True)

# ── 适配度图例 ────────────────────────────────────────
leg_row = len(data) + 5
ws1.merge_cells(f"A{leg_row}:H{leg_row}")
leg = ws1[f"A{leg_row}"]
leg.value = "★ 评分标准：5=优秀（强适配）｜4=良好｜3=一般（需二次开发）｜2=较弱｜1=差；价格区间仅供参考，实际以厂商正式报价为准"
leg.font  = Font(italic=True, color="595959", size=9, name="Arial")
leg.alignment = align(h="left", v="center")

# ══════════════════════════════════════════════════════
# SHEET 2 — 性价比综合评价
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet("性价比综合评价")
ws2.sheet_view.showGridLines = False

col_widths2 = [5, 18, 12, 12, 12, 12, 12, 12, 12, 12, 12, 15]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.row_dimensions[1].height = 18
ws2.row_dimensions[2].height = 36

ws2.merge_cells("A1:L1")
t2 = ws2["A1"]
t2.value = "MIM行业 ERP 性价比综合评价"
t2.font  = Font(bold=True, color=WHITE, size=14, name="Arial")
t2.fill  = fill(DARK_BLUE)
t2.alignment = align(h="center", v="center")

ws2.merge_cells("A2:L2")
sub2 = ws2["A2"]
sub2.value = "综合评分 = MIM适配度×35% + 功能完整度×25% + 价格友好度×20% + 实施便捷度×10% + 服务能力×10%"
sub2.font  = Font(color="595959", size=9, name="Arial", italic=True)
sub2.fill  = fill(GRAY_HDR)
sub2.alignment = align(h="center", v="center")

hdrs2 = ["排名", "品牌/产品", "MIM适配度\n(35%)", "功能完整度\n(25%)",
         "价格友好度\n(20%)", "实施便捷度\n(10%)", "服务能力\n(10%)", "综合得分", "价格区间",
         "适合企业规模", "推荐理由", "风险提示"]
for c, h in enumerate(hdrs2, 1):
    hdr(ws2, 3, c, h)

# 性价比数据（综合得分已按权重计算）
性价比_data = [
    (1,"鼎捷E10",      5,4.5,4.0,4.0,4.5, 4.43,"40-100万",  "年营收1-5亿中大型MIM","深耕制造业40年，MIM行业know-how积累最深厚，功能与价格平衡好","需评估渠道服务商质量"),
    (2,"鼎捷易飞",    4,4.0,4.5,5.0,4.0, 4.28,"15-40万",   "年营收<1亿中小MIM","15天极速上线，轻量化，适合MIM小批量多品种模式","大型项目功能深度不足"),
    (3,"用友U9 Cloud",4,4.5,3.5,4.0,4.5, 4.08,"25-80万/年","年营收5000万-5亿","财税合规强，插单处理灵活，政策更新快，适合内销型MIM","对MIM模具参数管理深度一般"),
    (4,"赛意信息SMOM",4,4.5,3.5,4.0,4.0, 4.03,"30-80万",   "精密制造/汽车零部件MIM","MES+ERP一体化，数据采集准确率99.8%，设备集成能力强","品牌知名度较低"),
    (5,"金蝶云·星空",3,4.0,4.5,4.0,3.5, 3.78,"8-25万/年", "年营收<2亿中小MIM","云原生轻量，盘点效率高60%，业财税一体，IT团队友好","MIM工艺路线管理深度不足"),
    (6,"华为云ERP",   4,4.0,3.5,4.0,4.5, 3.93,"30-100万/年","有国产化替代需求MIM","鲲鹏芯片+5G边缘计算，设备直连毫秒级，IoT融合深","价格透明度较低"),
    (7,"鼎捷T100",    5,5.0,2.0,3.0,5.0, 3.88,"100万+",    "年营收5亿+集团MIM","多工厂/跨国协同最优，数字孪生+MIM全链路，适合集团化布局","投入成本高，实施周期长"),
    (8,"浪潮GS Cloud",4,4.5,3.0,3.5,4.0, 3.73,"50-200万", "大型央国企背景MIM","联邦学习跨组织数据协同，财务管控强，适合多业态集团","用户体验一般"),
    (9,"用友BIP",     4,4.5,2.5,3.5,4.5, 3.63,"80万+/年",   "大型集团MIM企业","BIP平台生态完整，AI+RPA财务自动化97%，功能最全面","订阅费用高，SaaS迁移成本大"),
    (10,"SAP S/4HANA",5,5.0,1.5,2.5,4.5, 3.38,"80万+/模块\n千万级总投入","跨国大型MIM集团","全球合规最强，多工厂/多语言/多会计准则，适合出口型MIM","价格极高，本地化服务弱"),
    (11,"Oracle Fusion",5,5.0,1.5,2.5,4.0, 3.28,"80万+/模块","有全球布局的MIM企业","AI Agent原生，自动化率80%+，合规引擎覆盖200+国家","实施复杂，国产化适配差"),
    (12,"微软Dynamics",4,4.0,3.0,3.5,4.0, 3.58,"50-150万/年","已用Microsoft 365企业","与Office/Teams原生集成，Copilot加持，适合IT互联网企业","MIM行业know-how积累少"),
    (13,"金蝶云·星辰",2,3.0,5.0,4.5,3.0, 3.18,"数千元/年", "初创MIM小作坊","价格极低，上线快，适合业务刚起步的MIM工段","功能深度不足，不适合规模运营"),
    (14,"Infor ERP",  3,4.0,3.0,3.5,3.0, 3.18,"40-100万",  "机械/电子离散制造MIM","行业垂直模板预配置，实施较快","定制化灵活性低，国内服务商少"),
    (15,"QAD ERP",    3,4.5,2.5,3.0,3.0, 3.03,"50-120万",  "汽车/航空MIM精益生产","JIT/MTO精益管理，适合大批量稳定定单MIM","行业聚焦较窄，国内MIM适用性一般"),
]

for r_idx, row in enumerate(性价比_data, 4):
    ws2.row_dimensions[r_idx].height = 50
    bg_row = alt_rows[r_idx % 2]

    rank, brand = row[0], row[1]
    score = row[7]

    for c_idx, val in enumerate(row, 1):
        if c_idx == 1:  # 排名
            bg = "FFF2CC" if rank <= 3 else bg_row
            cell(ws2, r_idx, c_idx, rank, bold=True, bg=bg,
                 fg="595959" if rank > 3 else "7D6608", h="center")
        elif c_idx == 2:  # 品牌
            cell(ws2, r_idx, c_idx, brand, bold=True, bg=bg_row,
                 fg=DARK_BLUE, h="left")
        elif c_idx == 8:  # 综合得分
            bg, fg_c = score_color(round(score))
            c = cell(ws2, r_idx, c_idx, round(score, 2), bold=True,
                     bg=bg, fg=fg_c, h="center")
            c.number_format = "0.00"
        elif c_idx == 9:  # 价格区间
            cell(ws2, r_idx, c_idx, val, bg=bg_row, h="left", wrap=True)
        elif c_idx == 10:  # 规模
            cell(ws2, r_idx, c_idx, val, bg=bg_row, h="left", wrap=True)
        elif c_idx == 11:  # 推荐理由
            cell(ws2, r_idx, c_idx, val, bg=bg_row, h="left", wrap=True)
        elif c_idx == 12:  # 风险提示
            cell(ws2, r_idx, c_idx, val, bg="FFF0F0", h="left", wrap=True,
                 fg="9C0006")
        else:
            bg, fg_c = score_color(val)
            cell(ws2, r_idx, c_idx, val, bold=True, bg=bg, fg=fg_c,
                 h="center")

# ══════════════════════════════════════════════════════
# SHEET 3 — 选型决策指南
# ══════════════════════════════════════════════════════
ws3 = wb.create_sheet("选型决策指南")
ws3.sheet_view.showGridLines = False

for i, w in enumerate([5,20,25,25,22], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:E1")
t3 = ws3["A1"]
t3.value = "MIM行业 ERP 选型决策指南"
t3.font  = Font(bold=True, color=WHITE, size=14, name="Arial")
t3.fill  = fill(DARK_BLUE)
t3.alignment = align(h="center", v="center")
ws3.row_dimensions[1].height = 20

# 场景一：不同规模
scene_titles = [
    ("场景一：按企业规模选型", MID_BLUE, WHITE, 18),
    ("企业规模", "推荐产品", "核心原因", "预期投入", 22),
    ("初创/小微MIM\n（年营收<3000万）", "金蝶云·星辰\n鼎捷易飞", "价格极低，15-30天快速上线，基础功能覆盖销售/采购/库存/财务；金蝶操作体验好", "金蝶：数千元/年\n鼎捷易飞：15-20万（买断）", 50),
    ("中小型MIM\n（年营收3000万-1亿）", "鼎捷易飞\n金蝶云·星空", "鼎捷易飞：15天上线，MIM模具管理强；金蝶云星空：业财一体，条码集成效率提升60%", "易飞：15-40万\n金蝶星空：8-20万/年", 50),
    ("中大型MIM\n（年营收1-5亿）", "鼎捷E10\n用友U9 Cloud\n赛意SMOM", "鼎捷E10：制造业know-how最深；用友U9：财税合规+插单灵活；赛意SMOM：MES+ERP一体，数据采集99.8%", "鼎捷E10：40-100万\n用友U9：25-80万/年\n赛意：30-80万", 50),
    ("大型/集团MIM\n（年营收5亿+）", "鼎捷T100\nSAP S/4HANA\n用友BIP", "鼎捷T100：国产最推荐，多工厂/跨国合规；SAP：全球布局必备；用友BIP：AI+RPA最成熟", "鼎捷T100：100万+\nSAP：用友按项目\n用友BIP：80万+/年", 50),
    ("\n场景二：按业务痛点选型", MID_BLUE, WHITE, 18),
    ("核心痛点", "推荐产品", "解决方案亮点", "预期收益", 22),
    ("模具参数多\n工艺路线复杂", "鼎捷全系列\nQAD ERP", "鼎捷：模具BOP管理+工序级工艺路线+参数追溯；QAD：MTO按单生产最适合多工艺MIM", "工艺准备时间↓40%\n换线效率↑35%", 50),
    ("质量管控难\n追溯要求高", "鼎捷E10\n赛意SMOM\nOracle Fusion", "鼎捷：全流程质量追溯+批次管理；赛意：99.8%数据采集准确率；Oracle：工序级成本核算99.7%精度", "质量投诉↓60%\n追溯时间从小时→分钟", 50),
    ("设备集成难\nOT融合弱", "华为云ERP\n鼎捷E10\nSAP", "华为：鲲鹏+5G边缘计算，毫秒级设备直连；鼎捷：设备利用率98%；SAP：设备数字孪生成熟", "设备综合效率↑25%\n停机时间↓40%", 50),
    ("财税合规\n内控管理", "用友U9/BIP\n浪潮GS Cloud", "用友：金税四期完全合规，RPA发票审核自动化97%；浪潮：资金结算效率↑50%，合规预警92%准确率", "财务工时↓50%\n审计准备时间↓70%", 50),
    ("国产化替代\n政策要求", "华为云ERP\n鼎捷全系列\n浪潮GS Cloud", "鼎捷：国产化适配率100%（麒麟/统信/达梦）；华为：鲲鹏+欧拉全栈国产；浪潮：UBML低代码快速适配", "满足信创要求\n规避供应链风险", 50),
    ("出海/跨国\n多工厂协同", "SAP S/4HANA\nOracle Fusion\n鼎捷T100", "SAP：54种语言+200+会计准则自动切换；Oracle：AI Agent多国合规引擎；鼎捷T100：多工厂协同+国际合规", "全球可视化管理\n合规成本↓60%", 50),
]

row = 2
for item in scene_titles:
    if len(item) == 4:
        # 场景标题行
        title, bg_c, fg_c, height = item
        ws3.merge_cells(f"A{row}:E{row}")
        tc = ws3[f"A{row}"]
        tc.value = title
        tc.font  = Font(bold=True, color=fg_c, size=12, name="Arial")
        tc.fill  = fill(bg_c)
        tc.alignment = align(h="left", v="center")
        ws3.row_dimensions[row].height = height
        row += 1
    elif len(item) == 5:
        # 列标题行
        cols = item[:5]
        ws3.row_dimensions[row].height = 28
        for c, h in enumerate(cols, 1):
            hdr(ws3, row, c, h)
        row += 1
    else:
        # 数据行
        bg_row = alt_rows[row % 2]
        ws3.row_dimensions[row].height = item[-1]
        for c, val in enumerate(item[:5], 1):
            if c == 1:
                cell(ws3, row, c, val, bold=True, bg="FFF0F0" if "▼" in str(val) or "初创" in str(val) else bg_row,
                     fg=ACCENT_RED if c==1 else "000000", h="left", wrap=True)
            else:
                cell(ws3, row, c, val, bg=bg_row, h="left", wrap=True)
        row += 1

# ── 尾部免责 ──────────────────────────────────────────
row += 1
ws3.merge_cells(f"A{row}:E{row}")
disc = ws3[f"A{row}"]
disc.value = ("⚠️ 免责声明：以上内容基于2026年4月公开资料整理，不构成购买建议。"
              "ERP选型需结合企业实际业务、现有IT架构、团队能力和长期发展战略综合评估，"
              "建议安排厂商实地演示（POC）及参考同行实施案例后决策。")
disc.font  = Font(italic=True, color="7F7F7F", size=9, name="Arial")
disc.alignment = align(h="left", v="center", wrap=True)
ws3.row_dimensions[row].height = 36

# ── 冻结首行 ─────────────────────────────────────────
ws1.freeze_panes = "A4"
ws2.freeze_panes = "A4"
ws3.freeze_panes = "A3"

# ── 保存 ─────────────────────────────────────────────
out = "c:/Users/Administrator/WorkBuddy/Claw/MIM_ERP品牌对比.xlsx"
wb.save(out)
print(f"Saved: {out}")
