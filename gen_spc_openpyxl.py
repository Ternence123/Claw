"""
SPC 正态分布数据生成 + Excel 柱状图（含 USL/LSL/CL/UCL/LCL 标注）
使用 openpyxl 生成带嵌入式图表的 Excel 文件
"""

import openpyxl
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import math, random, datetime

# ── 参数 ──────────────────────────────────
NUM_GROUPS   = 25
SUBGROUP_SZ  = 5
MEAN         = 2.0000
SIGMA        = 0.003
USL          = 2.005
LSL          = 1.995
UCL          = MEAN + 3 * SIGMA   # 2.009
LCL          = MEAN - 3 * SIGMA   # 1.991

# ── Box-Muller 正态分布 ──────────────────
def randn():
    u = random.random()
    v = random.random()
    while u == 0: u = random.random()
    while v == 0: v = random.random()
    return math.sqrt(-2.0 * math.log(u)) * math.cos(2.0 * math.pi * v)

def normal(mean, sigma):
    return mean + sigma * randn()

# ── 生成数据 ──────────────────────────────
subgroups = []
all_data   = []

for i in range(NUM_GROUPS):
    samples = [round(normal(MEAN, SIGMA), 4) for _ in range(SUBGROUP_SZ)]
    all_data.extend(samples)
    avg_val  = round(sum(samples) / len(samples), 4)
    rng_val  = round(max(samples) - min(samples), 4)
    subgroups.append({
        'no':    i + 1,
        'samples': samples,
        'avg':  avg_val,
        'range':  rng_val,
    })

total_avg  = round(sum(all_data) / len(all_data), 4)
avg_range = round(sum(g['range'] for g in subgroups) / NUM_GROUPS, 4)

print(f"生成完成：{NUM_GROUPS} 组 × {SUBGROUP_SZ} = {len(all_data)} 个数据")
print(f"总均值 = {total_avg}, 平均极差 = {avg_range}")
print(f"UCL={UCL:.4f}, CL={MEAN}, LCL={LCL:.4f}")
print(f"USL={USL}, LSL={LSL}")

# ── 创建 Excel 工作簿 ────────────────────
wb = openpyxl.Workbook()
wb.properties.creator = "WorkBuddy"

# ── 工作表1：原始数据 ────────────────────
ws1 = wb.active
ws1.title = "原始数据"
headers = ["组号", "样本1", "样本2", "样本3", "样本4", "样本5", "组均值", "组极差"]
ws1.append(headers)

for g in subgroups:
    ws1.append([g['no']] + g['samples'] + [g['avg'], g['range']])

ws1.append(["汇总", "", "", "", "", "", total_avg, avg_range])

# 设置列宽
col_widths = [6, 10, 10, 10, 10, 10, 10, 10]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── 工作表2：直方图数据 + 柱状图 ─────────
ws2 = wb.create_sheet("SPC直方图")

# 计算直方图区间
min_v = min(all_data)
max_v = max(all_data)
bin_count = 20
bin_width = (max_v - min_v) / bin_count
bins = []
for i in range(bin_count):
    lo = min_v + i * bin_width
    hi = lo + bin_width
    bins.append({'lo': lo, 'hi': hi, 'count': 0, 'mid': (lo + hi) / 2})

for v in all_data:
    for b in bins:
        if b is bins[-1]:
            if b['lo'] <= v <= b['hi'] + 1e-9:
                b['count'] += 1
                break
        else:
            if b['lo'] <= v < b['hi']:
                b['count'] += 1
                break

# 写入直方图数据（含标注线列）
ws2.append(["区间中点", "频数", "USL", "LSL", "CL", "UCL", "LCL"])
for b in bins:
    ws2.append([
        round(b['mid'], 5),
        b['count'],
        USL,
        LSL,
        MEAN,
        round(UCL, 5),
        round(LCL, 5),
    ])

# 添加标注线说明
ws2.append([])
ws2.append(["标注", "数值", "说明"])
ws2.append(["USL(规格上限)", USL, "客户要求最大值"])
ws2.append(["LSL(规格下限)", LSL, "客户要求最小值"])
ws2.append(["CL(中心线)", MEAN, "过程总均值"])
ws2.append(["UCL(上控制限)", round(UCL, 5), "CL + 3σ"])
ws2.append(["LCL(下控制限)", round(LCL, 5), "CL - 3σ"])

# 设置列宽
for col, w in [(1, 12), (2, 8), (3, 10), (4, 10), (5, 10), (6, 10), (7, 10)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

# ── 创建柱状图 ────────────────────────────
chart = BarChart()
chart.title = "SPC 正态分布直方图（25组数据）"
chart.style = 10
chart.x_axis.title = "尺寸区间（mm）"
chart.y_axis.title = "频数（出现次数）"
chart.height = 12   # 图表高度（行数）
chart.width  = 24   # 图表宽度（列数）
chart.grouping = "standard"  # 标准柱状图

# 数据引用：频数（B列）
data_ref = Reference(ws2, min_col=2, min_row=1, max_row=bin_count+1)
cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=bin_count+1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

# 将图表插入到工作表中
ws2.add_chart(chart, "J2")
print("✅ 柱状图已添加到「SPC直方图」工作表")

# ── 工作表3：SPC 参数说明 ────────────────
ws3 = wb.create_sheet("SPC参数说明")
ws3.append(["SPC 关键参数", "数值", "计算公式", "说明"])
params = [
    ["过程均值 CL",  MEAN,     "—",        "目标值 / 总均值"],
    ["标准差 σ",    SIGMA,    "—",        "过程固有变异"],
    ["UCL(上控制限)", round(UCL,5), "CL + 3σ", "超出说明过程异常"],
    ["LCL(下控制限)", round(LCL,5), "CL - 3σ", "超出说明过程异常"],
    ["USL(规格上限)", USL,      "—",        "客户允许的最大值"],
    ["LSL(规格下限)", LSL,      "—",        "客户允许的最小值"],
]
for row in params:
    ws3.append(row)

ws3.append([])
ws3.append(["判断规则", "", "说明"])
ws3.append(["控制限 vs 规格限", "", "控制限是过程能力，规格限是客户要求，两者不同！"])
ws3.append(["超出控制限",       "", "说明过程异常，需查因"])
ws3.append(["超出规格限",       "", "说明产品不合格，需处置"])

for col, w in [(1, 20), (2, 18), (3, 18), (4, 35)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

# ── 输出文件 ──────────────────────────────
out_path = r"C:\Users\Administrator\Downloads\SPC正态分布数据_25组.xlsx"
wb.save(out_path)
print(f"\n✅ Excel 已保存：{out_path}")
print(f"   含 3 个工作表：原始数据 / SPC直方图 / SPC参数说明")
