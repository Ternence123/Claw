import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\Administrator\WorkBuddy\Claw\MIM_ERP品牌对比.xlsx')
print('Sheets:', wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'  {sn}: {ws.max_row} rows x {ws.max_column} cols')
print("Verification OK")
