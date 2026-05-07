from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

wb = Workbook()
ws = wb.active
ws.title = "物料流转流程"

# 配色方案
COLOR = {
    'primary': '065A82',      # 深蓝
    'secondary': '1C7293',    # 青蓝
    'accent': '21295C',       # 深夜蓝
    'light': 'CADCFC',        # 冰蓝
    'success': '10B981',      # 绿色
    'warning': 'F59E0B',      # 橙色
    'error': 'EF4444',        # 红色
    'bg': 'F8FAFC'            # 浅灰背景
}

# 定义边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 设置列宽
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12

# 标题行
ws['A1'] = '物料流转流程与岗位职责'
ws['A1'].font = Font(name='Microsoft YaHei', size=16, bold=True, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color=COLOR['primary'], end_color=COLOR['primary'], fill_type='solid')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A1:H1')
ws.row_dimensions[1].height = 30

# 表头
headers = ['序号', '流程步骤', '仓库', 'IQC', '生产', '生管', '成品委外仓库', '生产全检']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLOR['secondary'], end_color=COLOR['secondary'], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws.row_dimensions[2].height = 40

# 流程数据
process_data = [
    (1, '仓库原料收料', '●', '', '', '', '', ''),
    (2, 'IQC检验', '', '●', '', '', '', ''),
    (3, '仓库收料并对生产发料', '●', '', '', '', '', ''),
    (4, '生产制造', '', '', '●', '', '', ''),
    (5, '过程半成品检验', '', '', '■', '', '', ''),  # ■表示生产全检参与
    (6, '半成品入库', '●', '', '', '', '', ''),
    (7, '半成品外发加工', '', '', '', '●', '', ''),
    (8, '厂商加工', '', '', '', '●', '', ''),
    (9, '委外成品回厂', '', '', '', '', '●', ''),
    (10, '成品回厂检验', '', '■', '', '', '●', ''),  # ■表示IQC参与
    (11, '成品入库', '●', '', '', '', '', ''),
    (12, '成品出库全检', '●', '', '', '', '', '●'),
    (13, '成品全检入库', '●', '', '', '', '', '●'),
    (14, '成品检验', '', '■', '', '', '', '●'),  # ■表示IQC参与
    (15, '成品出货', '●', '', '', '●', '', '')  # 生管也参与
]

# 填充数据
for idx, (seq, step, wh, iqc, prod, pc, out_wh, prod_qc) in enumerate(process_data, 3):
    # 序号
    ws.cell(row=idx, column=1, value=seq).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=idx, column=1).border = thin_border
    ws.cell(row=idx, column=1).font = Font(name='Microsoft YaHei', size=10, bold=True)
    
    # 流程步骤
    ws.cell(row=idx, column=2, value=step).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.cell(row=idx, column=2).border = thin_border
    ws.cell(row=idx, column=2).font = Font(name='Microsoft YaHei', size=10)
    
    # 各岗位职责
    roles = [wh, iqc, prod, pc, out_wh, prod_qc]
    for col_offset, role in enumerate(roles, 3):
        cell = ws.cell(row=idx, column=col_offset, value=role if role else '')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.font = Font(name='Microsoft YaHei', size=14)
        
        # 设置颜色
        if role == '●':
            cell.font = Font(name='Microsoft YaHei', size=14, color=COLOR['primary'], bold=True)
            cell.fill = PatternFill(start_color=COLOR['light'], end_color=COLOR['light'], fill_type='solid')
        elif role == '■':
            cell.font = Font(name='Microsoft YaHei', size=14, color=COLOR['warning'], bold=True)
            cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    
    ws.row_dimensions[idx].height = 25

# 添加图例说明
legend_row = len(process_data) + 4
ws[f'A{legend_row}'] = '图例说明：'
ws[f'A{legend_row}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color=COLOR['primary'])

ws[f'A{legend_row+1}'] = '●'
ws[f'A{legend_row+1}'].font = Font(name='Microsoft YaHei', size=16, color=COLOR['primary'], bold=True)
ws[f'B{legend_row+1}'] = '主要责任'
ws[f'B{legend_row+1}'].font = Font(name='Microsoft YaHei', size=11)

ws[f'A{legend_row+2}'] = '■'
ws[f'A{legend_row+2}'].font = Font(name='Microsoft YaHei', size=16, color=COLOR['warning'], bold=True)
ws[f'B{legend_row+2}'] = '协作/参与'
ws[f'B{legend_row+2}'].font = Font(name='Microsoft YaHei', size=11)

# 添加流程说明
desc_row = legend_row + 4
ws[f'A{desc_row}'] = '流程说明：'
ws[f'A{desc_row}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color=COLOR['primary'])

descriptions = [
    '1. 仓库负责原料接收、存储和发放',
    '2. IQC负责进料检验和成品出货检验',
    '3. 生产部门负责制造和过程检验',
    '4. 生管负责生产计划和委外加工协调',
    '5. 成品委外仓库负责委外加工物料管理',
    '6. 生产全检负责成品出库和入库检验'
]

for i, desc in enumerate(descriptions, desc_row + 1):
    ws[f'A{i}'] = desc
    ws[f'A{i}'].font = Font(name='Microsoft YaHei', size=10)
    ws.merge_cells(f'A{i}:H{i}')

# 保存文件
output_file = '物料流转流程与岗位职责.xlsx'
wb.save(output_file)
print(f"✅ Excel文件生成成功：{output_file}")
print(f"📊 共{len(process_data)}个流程步骤")
print(f"👥 共6个岗位职责标注")
